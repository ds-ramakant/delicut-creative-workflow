import json
import os
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

IMAGES_DIR      = Path("outputs/images")
SESSION_LOG     = Path("outputs/session_log.xlsx")
PROMPTS_STAGING = Path("outputs/prompts_staging.json")

DIMENSIONS   = "1080x1080"
ASPECT_RATIO = "1:1"
ENGINE       = "Imagen4Ultra"
PERSONA      = "healthy-harry"

MODEL = "imagen-4.0-ultra-generate-001"

# Standard Gemini API client (not Vertex AI)
client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))


# ── Session ───────────────────────────────────────────────────────────────────

def make_session_id() -> str:
    now = datetime.now()
    return f"harry_{now.strftime('%m%d-%H%M')}"


def get_next_iteration() -> int:
    """Auto-detect the next iteration number from session_log.xlsx."""
    if not SESSION_LOG.exists():
        return 1
    wb = openpyxl.load_workbook(SESSION_LOG)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        iter_col = headers.index("iteration")
    except ValueError:
        return 1
    max_iter = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[iter_col]
        if isinstance(val, int) and val > max_iter:
            max_iter = val
    return max_iter + 1


# ── Prompts ───────────────────────────────────────────────────────────────────

def get_variants() -> list[dict]:
    """Read variants from prompts_staging.json — edit that file to change prompts."""
    if not PROMPTS_STAGING.exists():
        raise FileNotFoundError(f"Prompts staging file not found: {PROMPTS_STAGING}")
    data = json.load(PROMPTS_STAGING.open(encoding="utf-8"))
    return [
        {"variant": v["variant"], "slug": v["slug"], "prompt": v["prompt"]}
        for v in data["variants"]
    ]


# ── Image generation ──────────────────────────────────────────────────────────

def generate_image(prompt: str, output_path: Path) -> bool:
    """Generate an image using Imagen 4 Ultra via standard Gemini API."""
    response = client.models.generate_images(
        model=MODEL,
        prompt=prompt,
        config=types.GenerateImagesConfig(
            number_of_images=1,
            aspect_ratio=ASPECT_RATIO,
            person_generation="allow_adult",
        ),
    )

    if not response.generated_images:
        return False

    image_bytes = response.generated_images[0].image.image_bytes
    output_path.write_bytes(image_bytes)
    return True


# ── Excel helpers ─────────────────────────────────────────────────────────────

SESSION_HEADERS = [
    "session_id", "persona", "iteration", "variant",
    "engine", "dimensions", "prompt", "image_file",
    "score", "notes", "status",
]

HEADER_FILL  = PatternFill("solid", fgColor="043F12")
HEADER_FONT  = Font(bold=True, color="F9F4ED")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_headers(ws, headers: list[str]):
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22


def get_or_create_workbook(path: Path, headers: list[str]) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    _apply_headers(ws, headers)
    return wb


def append_row(ws, values: list):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Step 6: Generating images with Imagen 4 Ultra (Gemini API)\n")

    if not os.getenv("GOOGLE_API_KEY"):
        print("ERROR: GOOGLE_API_KEY not set in .env")
        return

    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    session_id = make_session_id()
    iteration  = get_next_iteration()
    variants   = get_variants()

    print(f"Session ID : {session_id}")
    print(f"Iteration  : {iteration}")
    print(f"Variants   : {len(variants)}")
    print(f"Engine     : {ENGINE}")
    print(f"Model      : {MODEL}")
    print(f"Dimensions : {DIMENSIONS}\n")

    wb_session = get_or_create_workbook(SESSION_LOG, SESSION_HEADERS)
    ws_session = wb_session.active

    for v in variants:
        label      = v["variant"]
        prompt     = v["prompt"]
        image_name = f"{session_id}_{v['slug']}_{ENGINE}.png"
        image_path = IMAGES_DIR / image_name

        print(f"[Variant {label}] {v['slug']} — generating...")

        try:
            success = generate_image(prompt, image_path)
            if success:
                print(f"  Saved  -> {image_path}")
                log_path = str(image_path)
            else:
                print(f"  WARNING: No image returned for variant {label}")
                log_path = "ERROR: no image returned"
        except Exception as e:
            print(f"  ERROR: {e}")
            log_path = f"ERROR: {e}"

        append_row(ws_session, [
            session_id, PERSONA, iteration, label,
            ENGINE, DIMENSIONS, prompt, log_path,
            "", "", "pending",
        ])

        if v != variants[-1]:
            time.sleep(1)

    SESSION_LOG.parent.mkdir(parents=True, exist_ok=True)
    wb_session.save(SESSION_LOG)

    print(f"\nDone.")
    print(f"Images saved to : {IMAGES_DIR}")
    print(f"Session log     : {SESSION_LOG}")
    print(f"\nNext: open session_log.xlsx, score each image (1-5),")
    print(f"add notes, set status to 'needs-refinement' or 'approved', save.")
    print(f"Then run step7_generate_adcopy.py for approved images.")


if __name__ == "__main__":
    main()
