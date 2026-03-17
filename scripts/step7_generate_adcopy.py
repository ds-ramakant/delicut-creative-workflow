import anthropic
import base64
import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from dotenv import load_dotenv

load_dotenv()

client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

SESSION_LOG = Path("outputs/session_log.xlsx")
ADCOPY_LOG  = Path("outputs/adcopy_log.xlsx")
COPY_DNA    = Path("outputs/copy_dna.json")

ADCOPY_HEADERS = [
    "session_id", "persona", "iteration", "variant",
    "engine", "dimensions", "image_file",
    "image_text", "ad_descriptors",
]

HEADER_FILL  = PatternFill("solid", fgColor="043F12")
HEADER_FONT  = Font(bold=True, color="F9F4ED")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

MEDIA_TYPES = {
    ".png":  "image/png",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
}

SYSTEM_PROMPT = """You are a world-class performance marketing copywriter for Delicut, a UAE-based healthy meal plan brand.

Brand identity:
- Colors: #043F12 (Spinach green), #F9F4ED (Cream), #FF3F1F (Grenade red), #EA5D29 (Pumpkin orange)
- Tone: Aspirational, direct, benefit-first. Never clinical or guilt-driven.
- Proven copy patterns from Delicut's top performers:
  - Two-tier structure: soft rhetorical question → bold punchy punchline (e.g. "COULD BE YOU!")
  - Second-person mirror: "COULD BE YOU" creates identification with the viewer
  - Rule-of-three benefits: "no cooking, no guilt, all gains"
  - Contrast: lowercase soft line → BOLD CAPS payoff
  - Identity-aligned promo code: BETTERYOU

Hard character limits (non-negotiable):
- Headline: maximum 30 characters including spaces
- Subline: maximum 90 characters including spaces"""


def load_copy_dna_context() -> str:
    """Load top-performer copy patterns to inspire generation."""
    if not COPY_DNA.exists():
        return ""
    data   = json.loads(COPY_DNA.read_text(encoding="utf-8"))
    images = data.get("images", [])
    # Collect up to 8 examples with real copy
    examples = []
    for img in images:
        vc = img.get("verbatim_copy", {})
        overlay = img.get("image_text_overlay")
        headline = vc.get("headline")
        if overlay or headline:
            examples.append({
                "image_text_overlay": overlay,
                "headline":           headline,
                "subline":            vc.get("subline"),
                "click_driver":       img.get("click_driver"),
                "emotional_angle":    img.get("emotional_angle"),
            })
        if len(examples) >= 8:
            break
    return json.dumps(examples, indent=2)


def read_approved_rows() -> list[dict]:
    if not SESSION_LOG.exists():
        return []
    wb      = openpyxl.load_workbook(SESSION_LOG)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]
    approved = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if str(row_dict.get("status", "")).strip().lower() == "approved":
            approved.append(row_dict)
    return approved


def generate_copy(image_path: Path, persona: str, copy_dna_context: str) -> dict:
    ext        = image_path.suffix.lower()
    media_type = MEDIA_TYPES.get(ext, "image/jpeg")
    image_data = base64.standard_b64encode(image_path.read_bytes()).decode("utf-8")

    dna_section = (
        f"\nTop-performer copy patterns for inspiration:\n{copy_dna_context}\n"
        if copy_dna_context else ""
    )

    prompt = f"""You are generating ad copy for an approved Delicut creative.

Persona: {persona}{dna_section}
STEP 1 — Read the image carefully. Identify:
- The subject, scene, or composition
- The dominant emotion or aspiration it communicates
- The single most striking visual moment

STEP 2 — Generate two copy outputs:

IMAGE TEXT (text overlay that appears directly ON the image):
- Inspired by what you see in the image — the copy must feel like a natural extension of the visual
- Short, bold, thumb-stopping — one or two lines maximum
- Follow Delicut's pattern: rhetorical question OR bold punchline, not both
- This is the first thing the viewer reads when they see the ad

AD DESCRIPTORS (copy outside the image frame, used by the designer):
- Headline: benefit-first, direct. MUST be ≤30 characters. Count every character including spaces.
- Subline: expands on the headline, persona-specific. MUST be ≤90 characters. Count every character including spaces.
- CTA: short action phrase (3–5 words)
- Offer: "25% Off your first plan | Use code: BETTERYOU"

Return ONLY a JSON object — no prose, no markdown fences:
{{
  "image_text": "...",
  "image_inspiration": "one sentence: what in the image drove this text",
  "headline": "...",
  "headline_char_count": 0,
  "subline": "...",
  "subline_char_count": 0,
  "cta": "...",
  "offer": "25% Off your first plan | Use code: BETTERYOU"
}}"""

    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": image_data},
                },
                {"type": "text", "text": prompt},
            ],
        }],
    )

    raw   = response.content[0].text.strip()
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            return {"error": "JSON parse failed", "raw": raw}
    return {"error": "No JSON found", "raw": raw}


def format_ad_descriptors(copy: dict) -> str:
    """Format headline/subline/CTA/offer into a single structured cell value."""
    headline = copy.get("headline", "")
    subline  = copy.get("subline", "")
    cta      = copy.get("cta", "")
    offer    = copy.get("offer", "")
    h_count  = copy.get("headline_char_count", len(headline))
    s_count  = copy.get("subline_char_count", len(subline))
    return (
        f"Headline [{h_count} chars]: {headline}\n"
        f"Subline [{s_count} chars]: {subline}\n"
        f"CTA: {cta}\n"
        f"Offer: {offer}"
    )


def _apply_headers(ws, headers: list):
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22


def get_or_create_workbook(path: Path, headers: list) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    _apply_headers(ws, headers)
    return wb


def already_logged(ws, image_file: str) -> bool:
    """Skip rows already written to adcopy_log (image_file is column 7)."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6] == image_file:
            return True
    return False


def main():
    print("Step 7: Generating ad copy for approved images\n")

    approved = read_approved_rows()
    if not approved:
        print("No approved rows in session_log.xlsx.")
        print("Score images and set status = 'approved' first, then re-run.")
        return

    print(f"Found {len(approved)} approved image(s)")

    copy_dna_context = load_copy_dna_context()
    if copy_dna_context:
        print("Copy DNA loaded from outputs/copy_dna.json")
    else:
        print("WARNING: copy_dna.json not found — run step3b first for richer copy patterns.\n")

    wb = get_or_create_workbook(ADCOPY_LOG, ADCOPY_HEADERS)
    ws = wb.active

    written = 0
    for row in approved:
        image_file = str(row.get("image_file", ""))
        image_path = Path(image_file)

        if not image_path.exists():
            print(f"  SKIP: image not found — {image_file}")
            continue

        if already_logged(ws, image_file):
            print(f"  SKIP: already in adcopy_log — {image_path.name}")
            continue

        persona = str(row.get("persona", ""))
        print(f"  [{row.get('variant')}] {image_path.name}")

        copy = generate_copy(image_path, persona, copy_dna_context)

        if "error" in copy:
            print(f"    ERROR: {copy['error']}")
            continue

        # Validate character limits — warn but don't block
        h_count = copy.get("headline_char_count", len(copy.get("headline", "")))
        s_count = copy.get("subline_char_count",  len(copy.get("subline", "")))
        if h_count > 30:
            print(f"    WARN: headline {h_count} chars — exceeds 30 limit")
        if s_count > 90:
            print(f"    WARN: subline {s_count} chars — exceeds 90 limit")

        ad_descriptors = format_ad_descriptors(copy)

        ws.append([
            row.get("session_id"),
            row.get("persona"),
            row.get("iteration"),
            row.get("variant"),
            row.get("engine"),
            row.get("dimensions"),
            image_file,
            copy.get("image_text", ""),
            ad_descriptors,
        ])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        print(f"    image_text    : {copy.get('image_text')}")
        print(f"    inspiration   : {copy.get('image_inspiration')}")
        print(f"    headline      : {copy.get('headline')}  ({h_count} chars)")
        print(f"    subline       : {copy.get('subline')}  ({s_count} chars)")
        print(f"    cta           : {copy.get('cta')}")
        written += 1

    ADCOPY_LOG.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ADCOPY_LOG)
    print(f"\nDone. {written} row(s) written to {ADCOPY_LOG}")


if __name__ == "__main__":
    main()
