import os
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

IMAGES_DIR  = Path("outputs/images")
SESSION_LOG = Path("outputs/session_log.xlsx")
DIMENSIONS  = "1080x1080"
ASPECT_RATIO = "1:1"
ENGINE       = "Imagen4Ultra"
PERSONA      = "healthy-harry"

client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))

# ── Session ───────────────────────────────────────────────────────────────────

def make_session_id() -> str:
    now = datetime.now()
    return f"harry_{now.strftime('%m%d-%H%M')}"

# ── Technical tail ────────────────────────────────────────────────────────────
# Appended to every variant prompt automatically.
# Covers camera system, settings, lighting rules, color treatment, and realism.

TECHNICAL_TAIL = """
CAMERA SYSTEM:
Canon R5 / Sony A7R IV. Natural lens perspective equivalent to 35mm–50mm full-frame.
No wide-angle distortion. No telephoto compression.

CAMERA SETTINGS (lifestyle / action):
ISO 400. Shutter 1/1600 sec. Aperture f/4. Color temperature 5600K, neutral white balance.

CAMERA SETTINGS (product flat-lay):
ISO 200. Shutter 1/160 sec. Aperture f/2.8. Color temperature 5200K, neutral white balance.

LIGHTING:
Natural daylight only. No artificial flash, no studio strobes, no colored gels.
Direction must be specified per scene. Quality defined as soft diffused or hard directional.
Shadows are soft falloff or crisp depending on scene — described per variant above.
Highlights are clean, not blown.

COLOR TREATMENT:
True-to-life color. Neutral whites — no yellow, orange, or blue cast.
Balanced contrast — never cinematic or overgraded. No LUTs. No filters. No stylization.

TEXTURE REALISM:
Visible skin texture — natural pores, light sweat sheen where appropriate.
Fabric shows natural behavior — wrinkles, stretch, folds at joints.
Product surfaces show material texture — matte plastic, fabric weave, metal finish.

DEPTH OF FIELD:
Subject or hero product in sharp focus. Background softly blurred but contextually readable.

COMPOSITION:
Intentional but unstyled. Everything feels observed, not staged.
Clear subject hierarchy. Use of negative space for premium, clean feel.
No floating objects. Correct gravity, shadows, and light falloff.

STRICT OUTPUT:
No text overlays. No logos other than Delicut branding on product.
No UI elements. No signage. Correct human anatomy and natural proportions.
Ultra-high-definition editorial realism. Natural grain, not noise-free plastic look.
"""


def build_prompt(scene_brief: str) -> str:
    """Combine the scene-specific brief with the universal technical tail."""
    return scene_brief.strip() + "\n" + TECHNICAL_TAIL.strip()


# ── Variants ──────────────────────────────────────────────────────────────────

def get_variants(session_id: str) -> list[dict]:
    return [
        {
            "variant": "A",
            "slug": "treadmill-female",
            "prompt": build_prompt("""
SCENE:
Modern commercial gym interior. Rows of treadmills, mirrors along one wall.
Clean, well-lit space. Natural light from floor-to-ceiling windows on the left side.
Background gym equipment and mirrors visible but softly out of focus.

SUBJECT:
Lean athletic female, mid-20s. Toned physique, natural proportions.
Wearing a fitted sports bra and high-waist training shorts, midriff visible.
Hair tied back in a high ponytail. Skin shows natural light sweat sheen.

POSE:
Running powerfully on a treadmill. Mid-stride, legs driving forward.
Strong forward gaze — candid intensity, not posed. Slight motion blur on legs suggesting real speed.
Natural weight distribution, forward lean of an actual runner.

LIGHTING:
Natural daylight from left-side floor-to-ceiling windows. Soft diffused light.
Soft shadow falloff on the right side of the body. Clean balanced exposure.
No overhead spot. No rim light.

CAMERA:
50mm equivalent. Eye-level angle, slightly front-side of the treadmill.
Mid-thigh to head crop. Background softly blurred, gym context still readable.
Use lifestyle/action camera settings.

MOOD:
Powerful, driven, in-the-zone. Energy of someone who makes no excuses.
"""),
        },
        {
            "variant": "B",
            "slug": "abstract-green",
            "prompt": build_prompt("""
SCENE:
Perfect overhead flat-lay on a deep spinach green surface, exact hex #043F12.
Clean, minimal studio setup. Generous negative space. No background clutter.

SUBJECT (HERO PRODUCT):
Delicut 900ml meal prep tray — primary and dominant focus of the frame.
Square form factor with soft rounded corners. Warm beige matte base.
Clear rigid plastic lid. Bold red wraparound label band with '/delicut/' in white lettering.

PRODUCT RULES:
Do not alter, restyle, recolor, or redesign the tray. Maintain exact shape, material, and branding.
Label must be clearly visible and readable.

ARRANGEMENT:
Tray centered in frame. At 10 o'clock: a single chalk-dusted white weight plate.
At 4 o'clock: a neatly folded white gym towel. At 7 o'clock: a matte stainless steel water bottle.
Each object naturally grounded with a realistic shadow beneath it.

LIGHTING:
Top-down overhead studio light. Hard directional, creating crisp clean shadows.
No fill light. No colored gels. Color temperature 5200K.
Use product flat-lay camera settings.

CAMERA:
50mm equivalent. Perfect 90-degree overhead angle, no perspective tilt.
Objects fill approximately 70% of frame. Negative space is the green surface.

MOOD:
Clean, precise, performance-driven. Gym gear meets macro-perfect meal prep.
"""),
        },
        {
            "variant": "C",
            "slug": "fridge-moment",
            "prompt": build_prompt("""
SCENE:
Modern kitchen. Clean, minimal interior — white or light stone countertop visible.
Refrigerator open in the background, interior light on. Time of day: morning.
Natural light from a window to the left, soft and warm.

SUBJECT:
Athletic male, late 20s to early 30s. Lean but not bulky — fitness-conscious physique.
Wearing a clean fitted white or grey t-shirt and training shorts.
Natural, relaxed posture — not flexing, not posed.

POSE:
Reaching into the fridge or holding a Delicut 900ml meal prep tray at chest height.
Looking at the tray with a natural satisfied expression — mid-moment, not smiling for camera.
Weight shifted slightly to one leg. Arm extended naturally toward fridge.

PRODUCT:
Delicut 900ml meal prep tray. Warm beige matte base, clear rigid lid,
bold red wraparound label reading '/delicut/' in white. Held naturally in one hand.
Slight tilt from the grip — not perfectly horizontal. Fingers wrapped naturally around base.

PRODUCT RULES:
Do not alter tray design, color, or label. Label must be clearly readable.

LIGHTING:
Natural morning light from left window. Soft diffused quality. 5600K neutral white.
Fridge interior light adds subtle warm fill from behind. No flash. No overhead spot.
Soft shadow on right side of subject.

CAMERA:
35mm equivalent. Eye-level, slightly elevated angle.
Three-quarter body shot — top of head to mid-thigh. Subject slightly left of center.
Background kitchen and open fridge softly out of focus but readable.
Use lifestyle/action camera settings.

MOOD:
Effortless healthy living. The fridge-smile moment — satisfaction without effort.
"""),
        },
        {
            "variant": "D",
            "slug": "post-workout-bench",
            "prompt": build_prompt("""
SCENE:
Gym locker room or quiet gym corner. Wooden bench in foreground.
Lockers or a plain gym wall softly out of focus in background.
Subdued, post-workout atmosphere — natural light from a high side window.

SUBJECT:
Athletic male, 28–32. Lean defined physique — not bulky, functionally fit.
Wearing dark fitted compression shorts and a sleeveless training top.
Natural post-workout state: slightly flushed, light sweat on arms and neck.

POSE:
Sitting on the bench. Forearms resting on knees, slight forward lean.
Head slightly down, composed and focused — the stillness after a hard session.
Not looking at camera. Hands relaxed, fingers loosely interlaced.

PRODUCT (PRIMARY FOCUS):
Delicut 900ml meal prep tray placed on the bench directly beside him — in sharp focus.
Warm beige matte base, clear rigid lid, bold red '/delicut/' label clearly visible.
Bag or gym kit partially visible beside tray, naturally suggesting he brought it for recovery.

PRODUCT RULES:
Tray is the sharpest element in the frame. Do not alter design, color, or label.
Label must be clearly readable.

LIGHTING:
Natural side light from a high window. Single light source, slightly directional.
Soft shadows across the locker room floor. Highlights on tray lid and subject's arms.
No artificial fill. No rim light. 5600K neutral white.

CAMERA:
35mm equivalent. Low to mid angle — camera slightly below bench height.
Focus on tray (sharp). Subject mid-ground, slightly blurred (depth of field).
Mid-shot — bench surface to subject's shoulders. Tray occupies left third of frame.
Use lifestyle/action camera settings.

MOOD:
Disciplined. Post-effort calm. The meal is the reward.
"""),
        },
    ]

# ── Excel helpers ─────────────────────────────────────────────────────────────

SESSION_HEADERS = [
    "session_id", "persona", "iteration", "variant",
    "engine", "dimensions", "prompt", "image_file",
    "score", "notes", "status",
]

HEADER_FILL  = PatternFill("solid", fgColor="043F12")
HEADER_FONT  = Font(bold=True, color="F9F4ED")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_headers(ws, headers: list[str]):
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22


def get_or_create_workbook(path: Path, headers: list[str]) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    _apply_headers(ws, headers)
    return wb


def append_row(ws, values: list):
    ws.append(values)
    # wrap text in long cells
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Image generation ──────────────────────────────────────────────────────────

def generate_image(prompt: str, output_path: Path) -> bool:
    response = client.models.generate_images(
        model="imagen-4.0-ultra-generate-001",
        prompt=prompt,
        config=types.GenerateImagesConfig(
            number_of_images=1,
            aspect_ratio=ASPECT_RATIO,
            person_generation="allow_adult",
        ),
    )
    if not response.generated_images:
        return False
    output_path.write_bytes(response.generated_images[0].image.image_bytes)
    return True


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Step 6: Generating images with Gemini Imagen\n")

    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    session_id = make_session_id()
    iteration  = 3
    variants   = get_variants(session_id)

    print(f"Session ID : {session_id}")
    print(f"Variants   : {len(variants)}")
    print(f"Engine     : {ENGINE}")
    print(f"Dimensions : {DIMENSIONS}\n")

    # Open or create session log
    wb_session = get_or_create_workbook(SESSION_LOG, SESSION_HEADERS)
    ws_session = wb_session.active

    for v in variants:
        label      = v["variant"]
        prompt     = v["prompt"]
        image_name = f"{session_id}_{v['slug']}_{ENGINE}.png"
        image_path = IMAGES_DIR / image_name

        print(f"[Variant {label}] Generating...")

        try:
            success = generate_image(prompt, image_path)
            if success:
                print(f"  Saved → {image_path}")
            else:
                print(f"  WARNING: No image returned for variant {label}")
                image_path = Path("ERROR - no image returned")
        except Exception as e:
            print(f"  ERROR: {e}")
            image_path = Path(f"ERROR: {e}")

        # Write to session_log
        append_row(ws_session, [
            session_id, PERSONA, iteration, label,
            ENGINE, DIMENSIONS, prompt, str(image_path),
            "", "", "pending",
        ])

        if label != variants[-1]["variant"]:
            time.sleep(1)

    SESSION_LOG.parent.mkdir(parents=True, exist_ok=True)
    wb_session.save(SESSION_LOG)

    print(f"\nDone.")
    print(f"Images saved to : {IMAGES_DIR}")
    print(f"Session log     : {SESSION_LOG}")
    print(f"\nNext step: open session_log.xlsx, score each image (1–5),")
    print(f"add notes, set status to 'needs-refinement' or 'approved', save.")
    print(f"Then run step7_generate_adcopy.py to generate copy for approved images.")


if __name__ == "__main__":
    main()
