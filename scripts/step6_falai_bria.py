import base64
import json
import os
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

IMAGES_DIR      = Path("outputs/images")
SESSION_LOG     = Path("outputs/session_log.xlsx")
PROMPTS_STAGING = Path("outputs/prompts_staging.json")

DIMENSIONS       = "1080x1080"
ENGINE           = "BriaProductShot"
PERSONA          = "healthy-harry"

# Default reference — used only if a variant has no reference_image set
DEFAULT_REFERENCE = Path("ads/reference/product_refs/tray_900ml_side.png")

# fal.ai model endpoint
MODEL_ID  = "fal-ai/bria/product-shot"
FAL_URL   = f"https://fal.run/{MODEL_ID}"

# Queue endpoint for async fallback (fal.run times out on slow models)
FAL_QUEUE_URL = f"https://queue.fal.run/{MODEL_ID}"

POLL_INTERVAL = 3
POLL_TIMEOUT  = 300


# ── fal.ai helpers ────────────────────────────────────────────────────────────

def _headers() -> dict:
    key = os.getenv("FAL_KEY")
    if not key:
        raise RuntimeError("FAL_KEY not set in .env")
    return {
        "Authorization": f"Key {key}",
        "Content-Type": "application/json",
    }


def _image_to_data_uri(path: Path) -> str:
    """Encode a local image file as a base64 data URI for fal.ai input."""
    ext = path.suffix.lower().lstrip(".")
    mime = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"
    b64 = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def _submit_and_poll(payload: dict) -> dict:
    """
    Submit to the fal.ai queue and poll until done.
    Falls back to async queue if the sync endpoint times out.
    """
    headers = _headers()

    # Try synchronous endpoint first (waits up to ~55s)
    try:
        resp = requests.post(FAL_URL, headers=headers, json=payload, timeout=90)
        if resp.status_code == 200:
            return resp.json()
        # If sync times out or returns 408, fall through to queue
        if resp.status_code not in (408, 504):
            resp.raise_for_status()
    except requests.exceptions.Timeout:
        pass

    # Async queue
    submit_resp = requests.post(
        FAL_QUEUE_URL,
        headers={**headers, "Content-Type": "application/json"},
        json=payload,
        timeout=30,
    )
    submit_resp.raise_for_status()
    request_id = submit_resp.json()["request_id"]
    status_url = f"{FAL_QUEUE_URL}/requests/{request_id}"

    elapsed = 0
    while True:
        time.sleep(POLL_INTERVAL)
        elapsed += POLL_INTERVAL
        if elapsed > POLL_TIMEOUT:
            raise TimeoutError(f"fal.ai request timed out after {POLL_TIMEOUT}s")
        status_resp = requests.get(status_url, headers=headers, timeout=30)
        status_resp.raise_for_status()
        data = status_resp.json()
        status = data.get("status")
        print(f"    status: {status} ({elapsed}s)")
        if status == "COMPLETED":
            # Fetch the actual result
            result_resp = requests.get(f"{status_url}/response", headers=headers, timeout=30)
            result_resp.raise_for_status()
            return result_resp.json()
        if status in ("FAILED", "CANCELLED"):
            raise RuntimeError(f"fal.ai request {status}: {data.get('error', '')}")


# ── Image generation ──────────────────────────────────────────────────────────

def generate_image(scene_description: str, output_path: Path, reference_image: Path) -> bool:
    """
    Place the Delicut product reference into a generated scene using Bria Product Shot.
    The model handles product rendering from the reference image — the prompt
    describes the scene only, not the product.
    """
    if not reference_image.exists():
        raise FileNotFoundError(f"Product reference not found: {reference_image}")

    product_image_uri = _image_to_data_uri(reference_image)

    payload = {
        "image_url": product_image_uri,
        "scene_description": scene_description,
        "shot_size": [1080, 1080],
        "num_results": 1,
        "fast": False,
    }

    result = _submit_and_poll(payload)

    images = result.get("images") or result.get("image")
    if not images:
        return False

    image_entry = images[0] if isinstance(images, list) else images
    image_url   = image_entry.get("url") if isinstance(image_entry, dict) else str(image_entry)

    img_resp = requests.get(image_url, timeout=60)
    img_resp.raise_for_status()
    output_path.write_bytes(img_resp.content)
    return True


# ── Session ───────────────────────────────────────────────────────────────────

def make_session_id() -> str:
    now = datetime.now()
    return f"harry_{now.strftime('%m%d-%H%M')}"


def get_next_iteration() -> int:
    if not SESSION_LOG.exists():
        return 1
    wb = openpyxl.load_workbook(SESSION_LOG)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        iter_col = headers.index("iteration")
    except ValueError:
        return 1
    max_iter = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[iter_col]
        if isinstance(val, int) and val > max_iter:
            max_iter = val
    return max_iter + 1


# ── Prompts ───────────────────────────────────────────────────────────────────

def get_variants() -> list[dict]:
    if not PROMPTS_STAGING.exists():
        raise FileNotFoundError(f"Not found: {PROMPTS_STAGING}")
    data = json.load(PROMPTS_STAGING.open(encoding="utf-8"))
    return [
        {
            "variant": v["variant"],
            "slug": v["slug"],
            "prompt": v["prompt"],
            "reference_image": Path(v["reference_image"]) if v.get("reference_image") else DEFAULT_REFERENCE,
        }
        for v in data["variants"]
    ]


# ── Excel helpers ─────────────────────────────────────────────────────────────

SESSION_HEADERS = [
    "session_id", "persona", "iteration", "variant",
    "engine", "dimensions", "prompt", "image_file",
    "score", "notes", "status",
]

HEADER_FILL  = PatternFill("solid", fgColor="043F12")
HEADER_FONT  = Font(bold=True, color="F9F4ED")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_headers(ws, headers):
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22


def get_or_create_workbook(path: Path, headers: list) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    _apply_headers(ws, headers)
    return wb


def append_row(ws, values: list):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Step 6: Generating images with Bria Product Shot (fal.ai)\n")
    print(f"Default reference : {DEFAULT_REFERENCE}")
    print(f"Model             : {MODEL_ID}\n")

    if not os.getenv("FAL_KEY"):
        print("ERROR: FAL_KEY not set in .env")
        return

    if not DEFAULT_REFERENCE.exists():
        print(f"ERROR: Default reference image not found: {DEFAULT_REFERENCE}")
        return

    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    session_id = make_session_id()
    iteration  = get_next_iteration()
    variants   = get_variants()

    print(f"Session ID : {session_id}")
    print(f"Iteration  : {iteration}")
    print(f"Variants   : {len(variants)}")
    print(f"Engine     : {ENGINE}")
    print(f"Dimensions : {DIMENSIONS}\n")

    wb_session = get_or_create_workbook(SESSION_LOG, SESSION_HEADERS)
    ws_session = wb_session.active

    for v in variants:
        label      = v["variant"]
        prompt     = v["prompt"]
        image_name = f"{session_id}_{v['slug']}_{ENGINE}.png"
        image_path = IMAGES_DIR / image_name

        ref = v["reference_image"]
        print(f"[Variant {label}] {v['slug']} — generating...")
        print(f"  Reference : {ref}")

        try:
            generate_image(prompt, image_path, ref)
            print(f"  Saved  -> {image_path}")
            log_path = str(image_path)
        except Exception as e:
            print(f"  ERROR: {e}")
            log_path = f"ERROR: {e}"

        append_row(ws_session, [
            session_id, PERSONA, iteration, label,
            ENGINE, DIMENSIONS, prompt, log_path,
            "", "", "pending",
        ])

        if v != variants[-1]:
            time.sleep(2)

    SESSION_LOG.parent.mkdir(parents=True, exist_ok=True)
    wb_session.save(SESSION_LOG)

    print(f"\nDone.")
    print(f"Images saved to : {IMAGES_DIR}")
    print(f"Session log     : {SESSION_LOG}")
    print(f"\nNext: open session_log.xlsx, score each image (1-5),")
    print(f"add notes, set status to 'needs-refinement' or 'approved', save.")
    print(f"Then run step7_generate_adcopy.py for approved images.")


if __name__ == "__main__":
    main()
