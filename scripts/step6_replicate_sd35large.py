import json
import os
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

IMAGES_DIR      = Path("outputs/images")
SESSION_LOG     = Path("outputs/session_log.xlsx")
PROMPTS_STAGING = Path("outputs/prompts_staging.json")

DIMENSIONS   = "1080x1080"
ASPECT_RATIO = "1:1"
ENGINE       = "SD35Large"
PERSONA      = "healthy-harry"

MODEL = "stability-ai/stable-diffusion-3.5-large"

# Negative prompt — things to exclude from every generation
NEGATIVE_PROMPT = (
    "extra hands, extra fingers, deformed hands, mutated hands, extra limbs, wrong anatomy, "
    "floating objects, text overlays, watermark, logo alterations, "
    "beige tray, red tray, wrong product color, blurry product, unreadable label, "
    "cartoon, illustration, painting, sketch, render, CGI, anime, unrealistic, "
    "oversaturated, overexposed, harsh shadows, lens flare, vignette"
)

# ── Replicate API ─────────────────────────────────────────────────────────────

REPLICATE_API_BASE = "https://api.replicate.com/v1"
POLL_INTERVAL      = 3    # seconds between status checks
POLL_TIMEOUT       = 300  # seconds before giving up


def _headers() -> dict:
    token = os.getenv("REPLICATE_API_TOKEN")
    if not token:
        raise RuntimeError("REPLICATE_API_TOKEN not set in .env")
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Prefer": "wait",
    }


def generate_image(prompt: str, output_path: Path) -> bool:
    """Generate an image using SD 3.5 Large via Replicate REST API."""
    headers = _headers()

    payload = {
        "input": {
            "prompt": prompt,
            "negative_prompt": NEGATIVE_PROMPT,
            "aspect_ratio": ASPECT_RATIO,
            "output_format": "png",
            "output_quality": 90,
            "cfg_scale": 4.5,
            "steps": 40,
        }
    }

    # 1. Create prediction
    create_url = f"{REPLICATE_API_BASE}/models/{MODEL}/predictions"
    resp = requests.post(create_url, headers=headers, json=payload, timeout=90)
    resp.raise_for_status()
    prediction = resp.json()

    # 2. Poll until done
    prediction_url = (
        prediction.get("urls", {}).get("get")
        or f"{REPLICATE_API_BASE}/predictions/{prediction['id']}"
    )

    elapsed = 0
    while prediction.get("status") not in ("succeeded", "failed", "canceled"):
        time.sleep(POLL_INTERVAL)
        elapsed += POLL_INTERVAL
        if elapsed > POLL_TIMEOUT:
            raise TimeoutError(f"Prediction timed out after {POLL_TIMEOUT}s")
        poll_resp = requests.get(prediction_url, headers=headers, timeout=30)
        poll_resp.raise_for_status()
        prediction = poll_resp.json()
        print(f"    status: {prediction.get('status')} ({elapsed}s elapsed)")

    if prediction.get("status") != "succeeded":
        raise RuntimeError(f"Prediction failed: {prediction.get('error', 'unknown error')}")

    # 3. Download the image
    output = prediction.get("output")
    if not output:
        return False
    image_url = output[0] if isinstance(output, list) else output
    img_resp = requests.get(image_url, timeout=60)
    img_resp.raise_for_status()
    output_path.write_bytes(img_resp.content)
    return True


# ── Session ───────────────────────────────────────────────────────────────────

def make_session_id() -> str:
    now = datetime.now()
    return f"harry_{now.strftime('%m%d-%H%M')}"


def get_next_iteration() -> int:
    """Auto-detect the next iteration number from session_log.xlsx."""
    if not SESSION_LOG.exists():
        return 1
    wb = openpyxl.load_workbook(SESSION_LOG)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        iter_col = headers.index("iteration")
    except ValueError:
        return 1
    max_iter = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[iter_col]
        if isinstance(val, int) and val > max_iter:
            max_iter = val
    return max_iter + 1


# ── Prompts ───────────────────────────────────────────────────────────────────

def get_variants() -> list[dict]:
    """Read variants from prompts_staging.json — edit that file to change prompts."""
    if not PROMPTS_STAGING.exists():
        raise FileNotFoundError(f"Prompts staging file not found: {PROMPTS_STAGING}")
    data = json.load(PROMPTS_STAGING.open(encoding="utf-8"))
    return [
        {"variant": v["variant"], "slug": v["slug"], "prompt": v["prompt"]}
        for v in data["variants"]
    ]


# ── Excel helpers ─────────────────────────────────────────────────────────────

SESSION_HEADERS = [
    "session_id", "persona", "iteration", "variant",
    "engine", "dimensions", "prompt", "image_file",
    "score", "notes", "status",
]

HEADER_FILL  = PatternFill("solid", fgColor="043F12")
HEADER_FONT  = Font(bold=True, color="F9F4ED")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_headers(ws, headers: list[str]):
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22


def get_or_create_workbook(path: Path, headers: list[str]) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    _apply_headers(ws, headers)
    return wb


def append_row(ws, values: list):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Step 6: Generating images with Stable Diffusion 3.5 Large (Replicate)\n")

    if not os.getenv("REPLICATE_API_TOKEN"):
        print("ERROR: REPLICATE_API_TOKEN not set in .env")
        return

    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    session_id = make_session_id()
    iteration  = get_next_iteration()
    variants   = get_variants()

    print(f"Session ID : {session_id}")
    print(f"Iteration  : {iteration}")
    print(f"Variants   : {len(variants)}")
    print(f"Engine     : {ENGINE}")
    print(f"Model      : {MODEL}")
    print(f"Dimensions : {DIMENSIONS}\n")

    wb_session = get_or_create_workbook(SESSION_LOG, SESSION_HEADERS)
    ws_session = wb_session.active

    for v in variants:
        label      = v["variant"]
        prompt     = v["prompt"]
        image_name = f"{session_id}_{v['slug']}_{ENGINE}.png"
        image_path = IMAGES_DIR / image_name

        print(f"[Variant {label}] {v['slug']} — generating...")

        try:
            generate_image(prompt, image_path)
            print(f"  Saved  -> {image_path}")
            log_path = str(image_path)
        except Exception as e:
            print(f"  ERROR: {e}")
            log_path = f"ERROR: {e}"

        append_row(ws_session, [
            session_id, PERSONA, iteration, label,
            ENGINE, DIMENSIONS, prompt, log_path,
            "", "", "pending",
        ])

        if v != variants[-1]:
            time.sleep(2)

    SESSION_LOG.parent.mkdir(parents=True, exist_ok=True)
    wb_session.save(SESSION_LOG)

    print(f"\nDone.")
    print(f"Images saved to : {IMAGES_DIR}")
    print(f"Session log     : {SESSION_LOG}")
    print(f"\nNext: open session_log.xlsx, score each image (1-5),")
    print(f"add notes, set status to 'needs-refinement' or 'approved', save.")
    print(f"Then run step7_generate_adcopy.py for approved images.")


if __name__ == "__main__":
    main()
